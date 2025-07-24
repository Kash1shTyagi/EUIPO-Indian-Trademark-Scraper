import logging
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import time
import pandas as pd
import re
from selenium.common.exceptions import StaleElementReferenceException
import os
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

chrome_opts = Options()
chrome_opts.add_argument("--headless")  

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_opts)
driver.get("https://euipo.europa.eu/ec2/search/find?language=en&text=&niceClass=1&size=25&page=1&harmonised=true&searchMode=WORDSPREFIX&sortBy=relevance")

# Wait for modal and try to close it
time.sleep(1)
try:
    close_btn = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'btn-primary') and text()='Close']"))
    )
    close_btn.click()
    logging.info("Disclaimer modal closed automatically.")
except Exception as e:
    logging.info("No disclaimer modal to close or already closed.")

# Wait a moment for modal to disappear
time.sleep(1)
logging.info("Proceeding with extraction...")

# Select the "India (CGPDTM)" checkbox by id
try:
    india_checkbox = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, "id_officeIN"))
    )
    if not india_checkbox.is_selected():
        india_checkbox.click()
    logging.info("India (CGPDTM) checkbox selected.")
except Exception as e:
    logging.warning(f"Could not select India checkbox: {e}")


# --- Extract ONLY first 10 pages for each Nice Class ---
for nice_class in range(1, 11):  # 1 to 10 inclusive
    try:
        nice_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "niceClass"))
        )
        nice_input.clear()
        nice_input.send_keys(str(nice_class))
        logging.info(f"Set Nice Class to {nice_class}")
    except Exception as e:
        logging.warning(f"Could not set Nice Class: {e}")
        continue

    try:
        search_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "proceed"))
        )
        search_button.click()
        logging.info("Search button clicked.")
        time.sleep(2)  # Let the table reload
    except Exception as e:
        logging.warning(f"Could not click Search button: {e}")
        continue

    logging.info(f"Extraction started for Nice Class {nice_class}.")

    data = []
    current_page = 1
    max_pages = 10
    while current_page <= max_pages:
        try:
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.ID, "advancedsearch_table"))
            )
            WebDriverWait(driver, 20).until(
                EC.text_to_be_present_in_element(
                    (By.XPATH, "//p[starts-with(text(), 'Page ')]"),
                    f"Page {current_page} of"
                )
            )
            table = driver.find_element(By.ID, "advancedsearch_table")
            rows = table.find_elements(By.XPATH, ".//tbody/tr")
            for row in rows:
                try:
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if len(cells) >= 10:
                        # Check for tick image in Harmonised and CGPDTM columns
                        harmonised_html = cells[3].get_attribute("innerHTML")
                        cgpdtm_html = cells[4].get_attribute("innerHTML")
                        harmonised_val = "✓" if "/ec2/static/images/tick.png" in harmonised_html else cells[3].text.strip()
                        cgpdtm_val = "✓" if "/ec2/static/images/tick.png" in cgpdtm_html else cells[4].text.strip()
                        data.append({
                            "Class": cells[1].text.strip(),
                            "Term": cells[2].text.strip(),
                            "Harmonised": harmonised_val,
                            "CGPDTM": cgpdtm_val,
                            "Harm": cells[5].text.strip(),
                            "Nice": cells[6].text.strip(),
                            "IDli": cells[7].text.strip(),
                            "Grou": cells[8].text.strip(),
                            "MGS": cells[9].text.strip(),
                        })
                except StaleElementReferenceException:
                    logging.warning("Stale row detected, skipping row.")
            logging.info(f"Extracted {len(rows)} rows from page {current_page}. Total so far: {len(data)}")
        except Exception as e:
            logging.warning(f"Could not extract table data: {e}")
            # --- Retry logic ---
            time.sleep(5)
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "advancedsearch_table"))
                )
                table = driver.find_element(By.ID, "advancedsearch_table")
                rows = table.find_elements(By.XPATH, ".//tbody/tr")
                for row in rows:
                    try:
                        cells = row.find_elements(By.TAG_NAME, "td")
                        if len(cells) >= 10:
                            harmonised_html = cells[3].get_attribute("innerHTML")
                            cgpdtm_html = cells[4].get_attribute("innerHTML")
                            harmonised_val = "✓" if "/ec2/static/images/tick.png" in harmonised_html else cells[3].text.strip()
                            cgpdtm_val = "✓" if "/ec2/static/images/tick.png" in cgpdtm_html else cells[4].text.strip()
                            data.append({
                                "Class": cells[1].text.strip(),
                                "Term": cells[2].text.strip(),
                                "Harmonised": harmonised_val,
                                "CGPDTM": cgpdtm_val,
                                "Harm": cells[5].text.strip(),
                                "Nice": cells[6].text.strip(),
                                "IDli": cells[7].text.strip(),
                                "Grou": cells[8].text.strip(),
                                "MGS": cells[9].text.strip(),
                            })
                    except StaleElementReferenceException:
                        logging.warning("Stale row detected, skipping row.")
                logging.info(f"Retry: Extracted {len(rows)} rows from page {current_page}. Total so far: {len(data)}")
            except Exception as e2:
                logging.warning(f"Retry failed on page {current_page}: {e2}")
                break

        if current_page == max_pages:
            logging.info(f"Reached max page limit (10) for Nice Class {nice_class}. Extraction complete.")
            break

        # Try to click the Next button
        try:
            next_btn_span = driver.find_element(By.ID, "listSource_table_next")
            next_links = next_btn_span.find_elements(By.TAG_NAME, "a")
            if not next_links or "disabled" in next_btn_span.get_attribute("class"):
                logging.info("Next button disabled. Extraction complete.")
                break
            else:
                next_links[0].click()
                logging.info(f"Next page ({current_page + 1}) clicked.")
                current_page += 1
                time.sleep(2)  # Let the next page load
        except Exception as e:
            logging.info("No Next button found or unable to click. Extraction complete.")
            break

    # Save to Excel for this class
    if data:
        df = pd.DataFrame(data)
        filename = f"Indian_Trademark_Class{nice_class}.xlsx"
        df.to_excel(filename, index=False)
        logging.info(f"Data saved to {filename}")
    else:
        logging.warning(f"No data extracted for Nice Class {nice_class}.")

driver.quit()

directory = "."
column_widths = {
    "Class": 10,
    "Term": 120,
    "Harmonised": 15,
    "CGPDTM": 15,
    "Harm": 10,
    "Nice": 10,
    "IDli": 10,
    "Grou": 10,
    "MGS": 10,
}

for fname in os.listdir(directory):
    if fname.startswith("Indian_Trademark_Class") and fname.endswith(".xlsx"):
        path = os.path.join(directory, fname)
        wb = load_workbook(path)
        ws = wb.active
        # Set width for each column if present
        for col in ws.iter_cols(1, ws.max_column):
            header = col[0].value
            if header in column_widths:
                ws.column_dimensions[col[0].column_letter].width = column_widths[header]
        wb.save(path)
        print(f"Updated column widths in {fname}")